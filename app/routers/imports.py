"""
Smart import router — 2-step upload → map → import.
Supports CSV, XLSX (Excel), and Google Sheets CSV exports.
"""
import csv
import io
import uuid
from datetime import datetime, timezone, timedelta
from typing import Any

import openpyxl
from fastapi import APIRouter, Depends, File, Form, Request, UploadFile
from fastapi.responses import RedirectResponse, StreamingResponse
from sqlalchemy.orm import Session

from app.database import get_db
from app.deps import require_user
from app.flash import get_flash, flash
from app.models import User, Lead, Client, Activity
from app.models.activity import ActivityType
from app.models.client import ClientType
from app.models.lead import LeadStatus
from app.templating import templates

router = APIRouter()

# ── In-memory session store (single-process; fine for Render single-worker) ──
_sessions: dict[str, dict] = {}
_SESSION_TTL = timedelta(minutes=30)


def _put_session(entity: str, headers: list[str], rows: list[dict]) -> str:
    sid = str(uuid.uuid4())
    _sessions[sid] = {
        "entity": entity,
        "headers": headers,
        "rows": rows,
        "expires": datetime.now(timezone.utc) + _SESSION_TTL,
    }
    _evict_old_sessions()
    return sid


def _get_session(sid: str) -> dict | None:
    s = _sessions.get(sid)
    if not s or s["expires"] < datetime.now(timezone.utc):
        _sessions.pop(sid, None)
        return None
    return s


def _evict_old_sessions():
    now = datetime.now(timezone.utc)
    expired = [k for k, v in _sessions.items() if v["expires"] < now]
    for k in expired:
        del _sessions[k]


# ── CRM field definitions ─────────────────────────────────────────────────────

LEAD_FIELDS = [
    {"key": "first_name", "label": "First Name",    "required": True},
    {"key": "last_name",  "label": "Last Name",     "required": False},
    {"key": "job_title",  "label": "Job Title",     "required": False},
    {"key": "company",    "label": "Company",       "required": False},
    {"key": "email",      "label": "Email",         "required": False},
    {"key": "mobile",     "label": "Mobile",        "required": False},
    {"key": "phone",      "label": "Phone",         "required": False},
    {"key": "linkedin_url","label": "LinkedIn URL", "required": False},
    {"key": "city",       "label": "City",          "required": False},
    {"key": "state",      "label": "State",         "required": False},
    {"key": "country",    "label": "Country",       "required": False},
    {"key": "source",     "label": "Source",        "required": False},
    {"key": "notes",      "label": "Notes",         "required": False},
]

CLIENT_FIELDS = [
    {"key": "name",     "label": "Company Name", "required": True},
    {"key": "type",     "label": "Type (direct/msp/partner/other)", "required": False},
    {"key": "industry", "label": "Industry",     "required": False},
    {"key": "website",  "label": "Website",      "required": False},
    {"key": "notes",    "label": "Notes",        "required": False},
]

# Auto-mapping aliases: crm_field → [possible column names]
_LEAD_ALIASES: dict[str, list[str]] = {
    "first_name":  ["first name", "firstname", "first", "given name", "name", "full name",
                    "fullname", "contact name", "lead name", "person"],
    "last_name":   ["last name", "lastname", "last", "surname", "family name"],
    "job_title":   ["job title", "title", "position", "role", "designation"],
    "company":     ["company", "company name", "organization", "organisation",
                    "account", "account name", "firm", "employer"],
    "email":       ["email", "email address", "e-mail", "mail", "email id"],
    "mobile":      ["mobile", "mobile number", "cell", "cell phone", "cellphone",
                    "contact 1", "contact1", "mobile no", "mob"],
    "phone":       ["phone", "phone number", "telephone", "contact number", "ph", "office phone",
                    "contact 2", "contact2", "tel", "landline"],
    "linkedin_url":["linkedin", "linkedin url", "linkedin profile", "linkedin link"],
    "city":        ["city", "town"],
    "state":       ["state", "province", "region"],
    "country":     ["country", "nation"],
    "source":      ["source", "lead source", "channel", "origin", "how did you hear"],
    "notes":       ["notes", "note", "description", "remarks",
                    "date of outreach", "outreach date", "first contact date"],
}

_CLIENT_ALIASES: dict[str, list[str]] = {
    "name":     ["name", "company", "company name", "account name", "organization",
                 "organisation", "firm", "client name"],
    "type":     ["type", "client type", "account type", "category", "tier"],
    "industry": ["industry", "sector", "vertical", "domain"],
    "website":  ["website", "url", "web", "site", "domain", "homepage"],
    "notes":    ["notes", "note", "comments", "description", "remarks"],
}


def _auto_map(headers: list[str], aliases: dict[str, list[str]]) -> dict[str, str]:
    """Return {crm_field: matched_header} best-effort auto-mapping."""
    normalised = {h.strip().lower(): h for h in headers}
    mapping = {}
    for field, candidates in aliases.items():
        for candidate in candidates:
            if candidate in normalised:
                mapping[field] = normalised[candidate]
                break
    return mapping


# ── File parsing ──────────────────────────────────────────────────────────────

MAX_BYTES = 5 * 1024 * 1024  # 5 MB


async def _parse_file(file: UploadFile) -> tuple[list[str], list[dict[str, Any]]] | str:
    """Return (headers, rows) or an error string."""
    raw = await file.read()
    if len(raw) > MAX_BYTES:
        return "File too large (max 5 MB)."

    fname = (file.filename or "").lower()

    if fname.endswith(".xlsx") or fname.endswith(".xls") or \
            file.content_type in ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                  "application/vnd.ms-excel"):
        return _parse_excel(raw)
    else:
        return _parse_csv(raw)


def _parse_excel(raw: bytes) -> tuple[list[str], list[dict]] | str:
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception as e:
        return f"Could not read Excel file: {e}"

    if not rows:
        return "Excel file is empty."

    # Track valid header positions — empty cells anywhere in the header row
    # must be skipped on both header AND data rows or columns will shift.
    raw_headers = [str(c).strip() if c is not None else "" for c in rows[0]]
    valid_indices = [i for i, h in enumerate(raw_headers) if h]
    headers: list[str] = []
    seen: dict[str, int] = {}
    for i in valid_indices:
        h = raw_headers[i]
        if h in seen:
            seen[h] += 1
            h = f"{h} ({seen[h]})"
        else:
            seen[h] = 1
        headers.append(h)
    if not headers:
        return "First row has no headers."

    data_rows = []
    for row in rows[1:]:
        values_full = [str(c).strip() if c is not None else "" for c in row]
        values = [values_full[i] if i < len(values_full) else "" for i in valid_indices]
        if not any(values):
            continue
        data_rows.append(dict(zip(headers, values)))

    return headers, data_rows


def _parse_csv(raw: bytes) -> tuple[list[str], list[dict]] | str:
    try:
        text = raw.decode("utf-8-sig", errors="replace")
    except Exception:
        return "Could not decode file as text."

    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel

    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    if not reader.fieldnames:
        return "CSV has no headers."

    headers = [h.strip() for h in reader.fieldnames if h and h.strip()]
    rows = []
    for row in reader:
        cleaned = {k.strip(): (v or "").strip() for k, v in row.items() if k and k.strip()}
        if not any(cleaned.values()):
            continue
        rows.append(cleaned)

    return headers, rows


# ── Template downloads ────────────────────────────────────────────────────────

@router.get("/leads/template.csv")
def leads_template():
    lines = ["first_name,last_name,job_title,company,email,mobile,phone,linkedin_url,city,state,country,source,notes",
             "John,Smith,Hiring Manager,Acme Corp,john@acme.com,+1 555 0100,+1 555 0200,https://linkedin.com/in/john,New York,NY,USA,LinkedIn,Met at conference"]
    return StreamingResponse(io.BytesIO("\n".join(lines).encode()),
                             media_type="text/csv",
                             headers={"Content-Disposition": "attachment; filename=contacts_template.csv"})


@router.get("/clients/template.csv")
def clients_template():
    lines = ["name,type,industry,website,notes",
             "Acme Healthcare,direct,Healthcare,https://acme.com,Key account",
             "MedForce MSP,msp,Staffing,https://medforce.com,"]
    return StreamingResponse(io.BytesIO("\n".join(lines).encode()),
                             media_type="text/csv",
                             headers={"Content-Disposition": "attachment; filename=clients_template.csv"})


# ── STEP 1 — Upload pages ─────────────────────────────────────────────────────

@router.get("/leads/import")
def leads_upload_page(request: Request, user: User = Depends(require_user)):
    return templates.TemplateResponse(request, "imports/upload.html", {
        "user": user, "flash": get_flash(request),
        "entity": "leads", "entity_label": "Leads",
        "back_url": "/leads",
        "fields": LEAD_FIELDS,
    })


@router.get("/clients/import")
def clients_upload_page(request: Request, user: User = Depends(require_user)):
    return templates.TemplateResponse(request, "imports/upload.html", {
        "user": user, "flash": get_flash(request),
        "entity": "clients", "entity_label": "Clients",
        "back_url": "/clients",
        "fields": CLIENT_FIELDS,
    })


@router.post("/leads/import")
async def leads_upload(
    request: Request,
    file: UploadFile = File(...),
    user: User = Depends(require_user),
):
    result = await _parse_file(file)
    if isinstance(result, str):
        return templates.TemplateResponse(request, "imports/upload.html", {
            "user": user, "entity": "leads", "entity_label": "Leads",
            "back_url": "/leads", "fields": LEAD_FIELDS, "error": result,
        }, status_code=400)
    headers, rows = result
    sid = _put_session("leads", headers, rows)
    return RedirectResponse(f"/import/leads/map/{sid}", status_code=303)


@router.post("/clients/import")
async def clients_upload(
    request: Request,
    file: UploadFile = File(...),
    user: User = Depends(require_user),
):
    result = await _parse_file(file)
    if isinstance(result, str):
        return templates.TemplateResponse(request, "imports/upload.html", {
            "user": user, "entity": "clients", "entity_label": "Clients",
            "back_url": "/clients", "fields": CLIENT_FIELDS, "error": result,
        }, status_code=400)
    headers, rows = result
    sid = _put_session("clients", headers, rows)
    return RedirectResponse(f"/import/clients/map/{sid}", status_code=303)


# ── STEP 2 — Mapping pages ────────────────────────────────────────────────────

@router.get("/leads/map/{sid}")
def leads_map_page(sid: str, request: Request, user: User = Depends(require_user)):
    session = _get_session(sid)
    if not session:
        return flash(RedirectResponse("/import/leads/import", 303), "Session expired — please re-upload.", "error")
    auto = _auto_map(session["headers"], _LEAD_ALIASES)
    return templates.TemplateResponse(request, "imports/map.html", {
        "user": user, "sid": sid, "entity": "leads", "entity_label": "Leads",
        "fields": LEAD_FIELDS, "file_headers": session["headers"],
        "auto_map": auto, "preview": session["rows"][:3],
        "total_rows": len(session["rows"]),
        "back_url": "/import/leads/import",
        "confirm_url": f"/import/leads/confirm/{sid}",
    })


@router.get("/clients/map/{sid}")
def clients_map_page(sid: str, request: Request, user: User = Depends(require_user)):
    session = _get_session(sid)
    if not session:
        return flash(RedirectResponse("/import/clients/import", 303), "Session expired — please re-upload.", "error")
    auto = _auto_map(session["headers"], _CLIENT_ALIASES)
    return templates.TemplateResponse(request, "imports/map.html", {
        "user": user, "sid": sid, "entity": "clients", "entity_label": "Clients",
        "fields": CLIENT_FIELDS, "file_headers": session["headers"],
        "auto_map": auto, "preview": session["rows"][:3],
        "total_rows": len(session["rows"]),
        "back_url": "/import/clients/import",
        "confirm_url": f"/import/clients/confirm/{sid}",
    })


# ── STEP 3 — Confirm + import ─────────────────────────────────────────────────

@router.post("/leads/confirm/{sid}")
async def leads_confirm(
    sid: str, request: Request,
    user: User = Depends(require_user), db: Session = Depends(get_db),
):
    session = _get_session(sid)
    if not session:
        return flash(RedirectResponse("/import/leads/import", 303), "Session expired — please re-upload.", "error")
    form = await request.form()
    mapping = {f["key"]: form.get(f"map_{f['key']}", "") for f in LEAD_FIELDS}
    results = _import_leads(session["rows"], mapping, user, db)
    del _sessions[sid]
    return templates.TemplateResponse(request, "imports/results.html", {
        "user": user, "entity": "Leads", "back_url": "/leads",
        "imported": results["imported"], "skipped": results["skipped"],
    })


@router.post("/clients/confirm/{sid}")
async def clients_confirm(
    sid: str, request: Request,
    user: User = Depends(require_user), db: Session = Depends(get_db),
):
    session = _get_session(sid)
    if not session:
        return flash(RedirectResponse("/import/clients/import", 303), "Session expired — please re-upload.", "error")
    form = await request.form()
    mapping = {f["key"]: form.get(f"map_{f['key']}", "") for f in CLIENT_FIELDS}
    results = _import_clients(session["rows"], mapping, user, db)
    del _sessions[sid]
    return templates.TemplateResponse(request, "imports/results.html", {
        "user": user, "entity": "Clients", "back_url": "/clients",
        "imported": results["imported"], "skipped": results["skipped"],
    })


# ── Import logic ──────────────────────────────────────────────────────────────

def _get_val(row: dict, col: str) -> str:
    return row.get(col, "").strip() if col else ""


def _trunc(value: str, maxlen: int) -> str | None:
    """Trim a value and truncate to maxlen — never let user input blow up an INSERT."""
    if not value:
        return None
    v = value.strip()
    return v[:maxlen] if v else None


def _import_leads(rows: list[dict], mapping: dict, user: User, db: Session) -> dict:
    imported, skipped = [], []
    first_col = mapping.get("first_name", "")
    # Build lookup of existing companies once so each row doesn't hit the DB
    client_map = {c.name.lower(): c.id for c in db.query(Client).all()}
    # Email dedupe lookup — case-insensitive
    existing_emails = {
        e.lower(): lid
        for (lid, e) in db.query(Lead.id, Lead.email).filter(Lead.email.isnot(None)).all()
        if e
    }
    # Name+company fallback dedupe set
    existing_name_co = {
        (f.lower().strip(), (l or "").lower().strip(), (c or "").lower().strip())
        for (f, l, c) in db.query(Lead.first_name, Lead.last_name, Lead.company).all()
        if f
    }
    # Also dedupe within this batch — re-importing a CSV with internal duplicates
    seen_emails: set[str] = set()
    seen_name_co: set[tuple[str, str, str]] = set()

    for i, row in enumerate(rows, start=2):
        first_name_raw = _get_val(row, first_col)
        if not first_name_raw:
            skipped.append({"row": i, "data": row, "reason": "First name is empty"})
            continue
        company_val = _trunc(_get_val(row, mapping.get("company", "")), 255)
        email_val = _trunc(_get_val(row, mapping.get("email", "")), 255)
        last_val = _trunc(_get_val(row, mapping.get("last_name", "")), 100)
        client_id = client_map.get(company_val.lower()) if company_val else None

        # Duplicate detection — email is the strongest signal
        if email_val:
            ek = email_val.lower()
            if ek in existing_emails:
                skipped.append({"row": i, "data": row, "reason": f"email '{email_val}' already exists (lead #{existing_emails[ek]})"})
                continue
            if ek in seen_emails:
                skipped.append({"row": i, "data": row, "reason": f"email '{email_val}' appears earlier in this file"})
                continue
            seen_emails.add(ek)

        # Fallback: same first+last+company already in DB (only if all three present)
        if first_name_raw and last_val and company_val:
            key = (first_name_raw.lower().strip(), last_val.lower().strip(), company_val.lower().strip())
            if key in existing_name_co:
                skipped.append({"row": i, "data": row, "reason": f"'{first_name_raw} {last_val}' at '{company_val}' already exists"})
                continue
            if key in seen_name_co:
                skipped.append({"row": i, "data": row, "reason": f"'{first_name_raw} {last_val}' at '{company_val}' appears earlier in this file"})
                continue
            seen_name_co.add(key)

        lead = Lead(
            first_name=_trunc(first_name_raw, 100),
            last_name=last_val,
            job_title=_trunc(_get_val(row, mapping.get("job_title", "")), 255),
            company=company_val,
            client_id=client_id,
            email=email_val,
            mobile=_trunc(_get_val(row, mapping.get("mobile", "")), 50),
            phone=_trunc(_get_val(row, mapping.get("phone", "")), 50),
            linkedin_url=_trunc(_get_val(row, mapping.get("linkedin_url", "")), 500),
            city=_trunc(_get_val(row, mapping.get("city", "")), 100),
            state=_trunc(_get_val(row, mapping.get("state", "")), 100),
            country=_trunc(_get_val(row, mapping.get("country", "")), 100),
            source=_trunc(_get_val(row, mapping.get("source", "")), 100),
            notes=_get_val(row, mapping.get("notes", "")) or None,  # TEXT column, no limit
            status=LeadStatus.NEW,
            owner_id=user.id,
        )
        db.add(lead)
        imported.append({
            "row": i,
            "name": lead.first_name,
            "company": lead.company or "",
        })

    db.commit()

    # Log action
    if imported:
        from app.services.audit import log_action
        log_action(
            db, user, "import_leads", "import",
            details={"count": len(imported), "skipped": len(skipped)}
        )

    return {"imported": imported, "skipped": skipped}


VALID_CLIENT_TYPES = {t.value for t in ClientType}


def _import_clients(rows: list[dict], mapping: dict, user: User, db: Session) -> dict:
    imported, skipped = [], []
    name_col = mapping.get("name", "")
    # Pre-load all existing names lowercased so each row doesn't hit the DB,
    # and so we catch case differences ("Acme" vs "ACME").
    existing_names = {n.lower().strip() for (n,) in db.query(Client.name).all() if n}
    seen_names: set[str] = set()

    for i, row in enumerate(rows, start=2):
        name = _get_val(row, name_col)
        if not name:
            skipped.append({"row": i, "data": row, "reason": "Name is empty"})
            continue

        type_raw = _get_val(row, mapping.get("type", "")).lower() or "direct"
        if type_raw not in VALID_CLIENT_TYPES:
            type_raw = "direct"

        nk = name.lower().strip()
        if nk in existing_names:
            skipped.append({"row": i, "data": row, "reason": f"'{name}' already exists"})
            continue
        if nk in seen_names:
            skipped.append({"row": i, "data": row, "reason": f"'{name}' appears earlier in this file"})
            continue
        seen_names.add(nk)

        client = Client(
            name=_trunc(name, 255),
            type=ClientType(type_raw),
            industry=_trunc(_get_val(row, mapping.get("industry", "")), 100),
            website=_trunc(_get_val(row, mapping.get("website", "")), 255),
            notes=_get_val(row, mapping.get("notes", "")) or None,  # TEXT column
            owner_id=user.id,
        )
        db.add(client)
        imported.append({"row": i, "name": client.name, "type": type_raw})

    db.flush()
    # Retroactively link unlinked leads to any newly-imported company by name
    for entry in imported:
        new_id = db.query(Client.id).filter(Client.name == entry["name"]).scalar()
        if new_id:
            db.query(Lead).filter(
                Lead.client_id.is_(None),
                Lead.company.ilike(entry["name"]),
            ).update({"client_id": new_id}, synchronize_session=False)
    db.commit()
    return {"imported": imported, "skipped": skipped}


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL SYNC — multi-sheet upsert with follow-up activities
# ══════════════════════════════════════════════════════════════════════════════

SYNC_EXTRA_FIELDS = [
    {"key": "status",          "label": "Status",             "required": False, "sync_only": True},
    {"key": "follow_up_notes", "label": "Follow-up / Update", "required": False, "sync_only": True},
    {"key": "follow_up_date",  "label": "Follow-up Date",     "required": False, "sync_only": True},
]
SYNC_LEAD_FIELDS = LEAD_FIELDS + SYNC_EXTRA_FIELDS

_SYNC_ALIASES: dict[str, list[str]] = {
    **_LEAD_ALIASES,
    # override notes to exclude "comments" so it routes to follow_up_notes instead
    "notes": ["notes", "note", "description", "remarks",
              "date of outreach", "outreach date", "first contact date"],
    "status": ["status", "lead status", "stage", "pipeline stage", "disposition"],
    "follow_up_notes": [
        "follow up", "follow-up", "followup", "update", "updates", "last update",
        "follow up notes", "followup notes", "activity", "last contact",
        "outreach update", "progress",
        "comments", "comment",  # Charmi sheet: Comments column
    ],
    "follow_up_date": [
        "follow up date", "follow-up date", "followup date", "next contact",
        "next follow up", "next followup", "due date", "next action date",
        "callback date", "reminder date",
        "fu1", "fu 1", "fu-1", "follow up 1", "followup1",  # Charmi sheet: FU1 column
    ],
}

_STATUS_MAP: dict[str, LeadStatus] = {
    "new": LeadStatus.NEW,
    "contacted": LeadStatus.CONTACTED,
    "reached out": LeadStatus.CONTACTED,
    "in progress": LeadStatus.CONTACTED,
    "follow up": LeadStatus.CONTACTED,
    "follow-up": LeadStatus.CONTACTED,
    "interested": LeadStatus.QUALIFIED,
    "qualified": LeadStatus.QUALIFIED,
    "hot": LeadStatus.QUALIFIED,
    "disqualified": LeadStatus.DISQUALIFIED,
    "not interested": LeadStatus.DISQUALIFIED,
    "rejected": LeadStatus.DISQUALIFIED,
    "lost": LeadStatus.DISQUALIFIED,
    "converted": LeadStatus.CONVERTED,
    "won": LeadStatus.CONVERTED,
    "closed won": LeadStatus.CONVERTED,
}


def _parse_excel_all_sheets(raw: bytes) -> "dict[str, tuple[list[str], list[dict]]] | str":
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as e:
        return f"Could not read Excel file: {e}"
    result: dict[str, tuple[list, list]] = {}
    for name in wb.sheetnames:
        ws = wb[name]
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            continue
        raw_headers = [str(c).strip() if c is not None else "" for c in all_rows[0]]
        valid_idx = [i for i, h in enumerate(raw_headers) if h]
        headers = [raw_headers[i] for i in valid_idx]
        if not headers:
            continue
        data_rows = []
        for row in all_rows[1:]:
            vals_full = [str(c).strip() if c is not None else "" for c in row]
            vals = [vals_full[i] if i < len(vals_full) else "" for i in valid_idx]
            if any(vals):
                data_rows.append(dict(zip(headers, vals)))
        if data_rows:
            result[name] = (headers, data_rows)
    wb.close()
    return result if result else "No worksheets with data were found."


def _parse_sync_date(value: str) -> "datetime | None":
    if not value:
        return None
    v = value.strip()
    try:
        serial = float(v)
        if 1000 < serial < 100000:
            from openpyxl.utils.datetime import from_excel
            d = from_excel(serial)
            if d:
                return datetime(d.year, d.month, d.day, 9, 0, tzinfo=timezone.utc)
    except ValueError:
        pass
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y",
                "%d-%m-%Y", "%m-%d-%Y", "%d %b %Y", "%b %d, %Y", "%B %d, %Y"):
        try:
            dt = datetime.strptime(v, fmt)
            return datetime(dt.year, dt.month, dt.day, 9, 0, tzinfo=timezone.utc)
        except ValueError:
            continue
    from app.services.outreach_reminder import _extract_date_from_notes
    d = _extract_date_from_notes(v)
    if d:
        return datetime(d.year, d.month, d.day, 9, 0, tzinfo=timezone.utc)
    return None


def _note_exists(db: Session, lead_id: int, body: str) -> bool:
    """True if a NOTE activity with this exact body already exists for the lead."""
    return db.query(Activity).filter(
        Activity.lead_id == lead_id,
        Activity.type == ActivityType.NOTE,
        Activity.body == body,
    ).first() is not None


def _task_exists_for_date(db: Session, lead_id: int, due: "datetime") -> bool:
    """True if an open TASK on the same calendar day already exists for the lead."""
    from sqlalchemy import func
    return db.query(Activity).filter(
        Activity.lead_id == lead_id,
        Activity.type == ActivityType.TASK,
        Activity.completed == False,  # noqa: E712
        func.date(Activity.due_at) == due.date(),
    ).first() is not None


def _sync_leads_from_excel(rows: list[dict], mapping: dict, user: User, db: Session) -> dict:
    created, updated, activities_added, skipped = [], [], [], []
    now = datetime.now(timezone.utc)

    existing_by_email: dict[str, Any] = {
        lead.email.lower(): lead
        for lead in db.query(Lead).filter(Lead.email.isnot(None)).all()
    }
    client_map = {c.name.lower(): c.id for c in db.query(Client).all()}

    # scalar fields that can be overwritten if the sheet has a non-empty value
    _UPDATABLE = [
        ("first_name",   100),
        ("last_name",    100),
        ("job_title",    255),
        ("company",      255),
        ("mobile",        50),
        ("phone",         50),
        ("linkedin_url", 500),
        ("city",         100),
        ("state",        100),
        ("country",      100),
        ("source",       100),
    ]

    for i, row in enumerate(rows, start=2):
        first_name = _trunc(_get_val(row, mapping.get("first_name", "")), 100)
        if not first_name:
            skipped.append({"row": i, "reason": "First name is empty"})
            continue

        email_val = _trunc(_get_val(row, mapping.get("email", "")), 255)
        follow_up_notes = _get_val(row, mapping.get("follow_up_notes", ""))
        follow_up_date_str = _get_val(row, mapping.get("follow_up_date", ""))
        new_status = _STATUS_MAP.get(_get_val(row, mapping.get("status", "")).lower().strip())
        existing = existing_by_email.get(email_val.lower()) if email_val else None

        if existing:
            # Update all mapped scalar fields if the sheet has a non-empty value
            for field, maxlen in _UPDATABLE:
                val = _trunc(_get_val(row, mapping.get(field, "")), maxlen)
                if val:
                    setattr(existing, field, val)
                    # keep client_id in sync when company changes
                    if field == "company":
                        existing.client_id = client_map.get(val.lower())

            if new_status:
                existing.status = new_status

            raw_note = _get_val(row, mapping.get("notes", ""))
            if raw_note and raw_note not in (existing.notes or ""):
                existing.notes = ((existing.notes + "\n\n") if existing.notes else "") + raw_note
                existing.outreach_notes_hash = None
                # Classify notes using NLP
                from app.services.nlp_classifier import classify_notes
                existing.outreach_category = classify_notes(existing.notes)

            updated.append({"name": existing.name, "id": existing.id})

            # NOTE activity — skip if identical body already logged
            if follow_up_notes and not _note_exists(db, existing.id, follow_up_notes):
                db.add(Activity(
                    type=ActivityType.NOTE,
                    subject=follow_up_notes[:120],
                    body=follow_up_notes,
                    lead_id=existing.id,
                    client_id=existing.client_id,
                    created_by_id=user.id,
                    completed=True,
                    completed_at=now,
                ))
                activities_added.append({"name": existing.name, "kind": "note"})

            # TASK reminder — skip if an open task already exists on that date
            if follow_up_date_str:
                due = _parse_sync_date(follow_up_date_str)
                if due and due > now and not _task_exists_for_date(db, existing.id, due):
                    db.add(Activity(
                        type=ActivityType.TASK,
                        subject=f"Follow up: {existing.name}",
                        body=follow_up_notes or "Follow-up from Excel sync",
                        lead_id=existing.id,
                        client_id=existing.client_id,
                        created_by_id=user.id,
                        due_at=due,
                        completed=False,
                    ))
                    activities_added.append({"name": existing.name, "kind": "task"})

        else:
            company_val = _trunc(_get_val(row, mapping.get("company", "")), 255)
            client_id = client_map.get(company_val.lower()) if company_val else None
            notes_val = _get_val(row, mapping.get("notes", "")) or None
            if follow_up_notes:
                notes_val = ((notes_val + "\n\n") if notes_val else "") + follow_up_notes
            lead = Lead(
                first_name=first_name,
                last_name=_trunc(_get_val(row, mapping.get("last_name", "")), 100),
                job_title=_trunc(_get_val(row, mapping.get("job_title", "")), 255),
                company=company_val,
                client_id=client_id,
                email=email_val,
                mobile=_trunc(_get_val(row, mapping.get("mobile", "")), 50),
                phone=_trunc(_get_val(row, mapping.get("phone", "")), 50),
                linkedin_url=_trunc(_get_val(row, mapping.get("linkedin_url", "")), 500),
                city=_trunc(_get_val(row, mapping.get("city", "")), 100),
                state=_trunc(_get_val(row, mapping.get("state", "")), 100),
                country=_trunc(_get_val(row, mapping.get("country", "")), 100),
                source=_trunc(_get_val(row, mapping.get("source", "")), 100),
                notes=notes_val,
                status=new_status or LeadStatus.NEW,
                owner_id=user.id,
            )
            # Classify notes using NLP on creation
            if notes_val:
                from app.services.nlp_classifier import classify_notes
                lead.outreach_category = classify_notes(notes_val)
            db.add(lead)
            db.flush()
            if email_val:
                existing_by_email[email_val.lower()] = lead
            if follow_up_date_str:
                due = _parse_sync_date(follow_up_date_str)
                if due and due > now:
                    db.add(Activity(
                        type=ActivityType.TASK,
                        subject=f"Follow up: {lead.name}",
                        body=follow_up_notes or "Follow-up from Excel sync",
                        lead_id=lead.id,
                        client_id=lead.client_id,
                        created_by_id=user.id,
                        due_at=due,
                        completed=False,
                    ))
                    activities_added.append({"name": lead.name, "kind": "task"})
            created.append({"name": lead.name, "id": lead.id})

    db.commit()

    # Update caches and auto-transition status for all leads that got activities
    from app.services.lead_cache import update_next_follow_up_cache
    from app.services.status_transition import auto_transition_on_activity

    lead_ids_with_activities = set()
    for item in created + updated:
        lead_id = item.get("id")
        if lead_id:
            lead_ids_with_activities.add(lead_id)

    for lead_id in lead_ids_with_activities:
        update_next_follow_up_cache(db, lead_id)
        auto_transition_on_activity(db, lead_id)

    # Log action
    if created or updated:
        from app.services.audit import log_action
        log_action(
            db, user, "excel_sync", "import",
            details={"created": len(created), "updated": len(updated), "activities": len(activities_added), "skipped": len(skipped)}
        )

    return {"created": created, "updated": updated, "activities": activities_added, "skipped": skipped}


@router.get("/excel-sync")
def excel_sync_page(request: Request, user: User = Depends(require_user)):
    return templates.TemplateResponse(request, "imports/excel_sync_upload.html", {
        "user": user, "flash": get_flash(request),
    })


@router.post("/excel-sync")
async def excel_sync_upload(
    request: Request,
    file: UploadFile = File(...),
    user: User = Depends(require_user),
):
    raw = await file.read()
    if len(raw) > MAX_BYTES:
        return templates.TemplateResponse(request, "imports/excel_sync_upload.html", {
            "user": user, "error": "File too large (max 5 MB).",
        }, status_code=400)
    fname = (file.filename or "").lower()
    if not (fname.endswith(".xlsx") or fname.endswith(".xls")):
        return templates.TemplateResponse(request, "imports/excel_sync_upload.html", {
            "user": user, "error": "Only .xlsx / .xls files are supported for Excel Sync.",
        }, status_code=400)
    result = _parse_excel_all_sheets(raw)
    if isinstance(result, str):
        return templates.TemplateResponse(request, "imports/excel_sync_upload.html", {
            "user": user, "error": result,
        }, status_code=400)
    all_sheets_serial = {
        name: {"headers": headers, "rows": rows, "count": len(rows)}
        for name, (headers, rows) in result.items()
    }
    sid = str(uuid.uuid4())
    _sessions[sid] = {
        "entity": "excel_sync",
        "all_sheets": all_sheets_serial,
        "expires": datetime.now(timezone.utc) + _SESSION_TTL,
    }
    _evict_old_sessions()
    return RedirectResponse(f"/import/excel-sync/sheets/{sid}", status_code=303)


@router.get("/excel-sync/sheets/{sid}")
def excel_sync_sheets(sid: str, request: Request, user: User = Depends(require_user)):
    session = _get_session(sid)
    if not session:
        return flash(RedirectResponse("/import/excel-sync", 303), "Session expired.", "error")
    return templates.TemplateResponse(request, "imports/excel_sync_sheets.html", {
        "user": user, "sid": sid,
        "sheets": [
            {"name": name, "count": info["count"], "headers": info["headers"][:6]}
            for name, info in session["all_sheets"].items()
        ],
    })


@router.post("/excel-sync/sheets/{sid}")
async def excel_sync_sheets_post(sid: str, request: Request, user: User = Depends(require_user)):
    session = _get_session(sid)
    if not session:
        return flash(RedirectResponse("/import/excel-sync", 303), "Session expired.", "error")
    form = await request.form()
    selected = form.getlist("sheets")
    if not selected:
        return templates.TemplateResponse(request, "imports/excel_sync_sheets.html", {
            "user": user, "sid": sid,
            "sheets": [
                {"name": n, "count": i["count"], "headers": i["headers"][:6]}
                for n, i in session["all_sheets"].items()
            ],
            "error": "Select at least one worksheet.",
        })
    all_sheets = session["all_sheets"]
    merged_rows: list[dict] = []
    seen_headers: list[str] = []
    for name in selected:
        if name not in all_sheets:
            continue
        info = all_sheets[name]
        for h in info["headers"]:
            if h not in seen_headers:
                seen_headers.append(h)
        for row in info["rows"]:
            row["__sheet__"] = name
            merged_rows.append(row)
    session["selected"] = selected
    session["headers"] = seen_headers
    session["rows"] = merged_rows
    return RedirectResponse(f"/import/excel-sync/map/{sid}", status_code=303)


@router.get("/excel-sync/map/{sid}")
def excel_sync_map(sid: str, request: Request, user: User = Depends(require_user)):
    session = _get_session(sid)
    if not session:
        return flash(RedirectResponse("/import/excel-sync", 303), "Session expired.", "error")
    auto = _auto_map(session["headers"], _SYNC_ALIASES)
    return templates.TemplateResponse(request, "imports/excel_sync_map.html", {
        "user": user, "sid": sid,
        "fields": SYNC_LEAD_FIELDS,
        "file_headers": session["headers"],
        "auto_map": auto,
        "preview": session["rows"][:3],
        "total_rows": len(session["rows"]),
        "selected_sheets": session.get("selected", []),
    })


@router.post("/excel-sync/confirm/{sid}")
async def excel_sync_confirm(
    sid: str, request: Request,
    user: User = Depends(require_user), db: Session = Depends(get_db),
):
    session = _get_session(sid)
    if not session:
        return flash(RedirectResponse("/import/excel-sync", 303), "Session expired.", "error")
    form = await request.form()
    mapping = {f["key"]: form.get(f"map_{f['key']}", "") for f in SYNC_LEAD_FIELDS}
    results = _sync_leads_from_excel(session["rows"], mapping, user, db)
    del _sessions[sid]
    return templates.TemplateResponse(request, "imports/excel_sync_results.html", {
        "user": user,
        "created": results["created"],
        "updated": results["updated"],
        "activities": results["activities"],
        "skipped": results["skipped"],
    })
